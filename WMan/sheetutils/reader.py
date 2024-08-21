from openpyxl import load_workbook


class SheetReader:
    def __init__(self, filepath: str):
        self.workbook = load_workbook(filepath)
        if self.workbook.active:
            self.sheet = self.workbook.active

    def get_data(self, start_row: int = 2):
        read_data = [
            list(row)
            for row in self.sheet.iter_rows(values_only=True, min_row=start_row)
        ]
        return read_data
