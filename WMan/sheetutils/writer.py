from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


class SheetWriter:
    def __init__(self):
        self.workbook = Workbook()
        if self.workbook.active:
            self.sheet = self.workbook.active

    def add_row_index_column(self):
        self.sheet.insert_cols(1)
        _ = self.sheet.cell(row=1, column=1, value="Row index")

        for row_num in range(1, self.sheet.max_row):
            _ = self.sheet.cell(row=row_num + 1, column=1, value=str(row_num))

    def add_headers(self, headers: list[str]):
        self.sheet.insert_rows(1)

        for col_index, header in enumerate(headers):
            _ = self.sheet.cell(1, col_index + 1, value=header)

    def add_data(self, data: list[list[int | str]]) -> None:
        for row in data:
            self.sheet.append(row)

    def set_optimal_column_widths(self):
        for column in self.sheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = max_length + 2
            self.sheet.column_dimensions[
                get_column_letter(column[0].column)
            ].width = adjusted_width

    def make_table(self, table_name: str):
        center_aligned_text = NamedStyle(
            name="center_aligned_text",
            alignment=Alignment(horizontal="center", vertical="center"),
        )

        if not self.sheet.parent:
            raise Exception("Sheet parent doesn't exist'")
        self.sheet.parent.add_named_style(center_aligned_text)

        table = Table(displayName=table_name, ref=self.sheet.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        self.sheet.add_table(table)

        for row in self.sheet[table.ref]:
            for cell in row:
                cell.style = center_aligned_text

    def set_column_currency_format(self, column_index: int):
        for row in self.sheet.iter_rows(
            min_row=2, min_col=column_index, max_col=column_index
        ):
            for cell in row:
                cell.number_format = "#,##0_-[$ريال-fa-IR]"
    
    def add_header(self, header: str):
        header_style = NamedStyle("header_style", alignment=Alignment(horizontal="center", vertical="center"), font=Font(bold=True, size=20))
        self.sheet.insert_rows(1)
        self.sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=self.sheet.max_column)
        self.sheet.cell(1, 1, value=header)
        self.sheet.cell(1,1).style = header_style

    def add_subheader(self, left_header: str, right_header: str):
        header_style = NamedStyle("sub_header_style", alignment=Alignment(horizontal="center", vertical="center"), font=Font(bold=True, size=10))
        self.sheet.insert_rows(1)
        self.sheet.cell(1, 1, value=left_header)
        self.sheet.cell(1, self.sheet.max_column, value=right_header)
        self.sheet.cell(1, 1).style = header_style
        self.sheet.cell(1, self.sheet.max_column).style = header_style

        

    def save(self, filename: str):
        self.workbook.save(filename)


if __name__ == "__main__":
    # Generate a large dataset
    num_rows = 1000
    num_cols = 50
    sample_data = [
        [f"Data {row}-{col}" for col in range(1, num_cols + 1)]
        for row in range(1, num_rows + 1)
    ]

    writer = SheetWriter()

    # Add data, headers, row index column, and make it a table
    writer.add_data(sample_data)
    writer.add_headers([f"Col {i}" for i in range(1, num_cols + 1)])
    writer.add_row_index_column()
    writer.make_table("huge_table")
    writer.set_optimal_column_widths()
    writer.add_subheader("Left Header", "Right Header")
    writer.add_header("Huge Table")

    writer.save(filename="../../sample.xlsx")

    # Generate a car sample pricelist:

    # data = [["Mercedes", 7123472743274], ["BMW", 7123471724], ["Lamborghini", 366246612]]
    # headers = ["Brand", "Price"]
    #
    # writer = SheetWriter()
    # writer.add_data(data)
    # writer.add_headers(headers)
    # writer.add_row_index_column()
    # writer.make_table("Car_Prices")
    # writer.set_column_currency_format(3)
    # writer.set_optimal_column_widths()
    # writer.save(filename="sample.xlsx")
