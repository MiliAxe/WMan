from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from typing import List


def add_row_index_column(sheet: Worksheet):
    sheet.insert_cols(1)
    sheet.cell(row=1, column=1, value="Row index")

    for row_num in range(1, sheet.max_row):
        sheet.cell(row=row_num + 1, column=1, value=str(row_num))


def add_headers(sheet: Worksheet, headers: List[str]):
    sheet.insert_rows(1)

    for col_index, header in enumerate(headers):
        sheet.cell(1, col_index + 1, value=header)


def add_data_to_sheet(worksheet: Worksheet, data: List[List]) -> None:
    for row in data:
        worksheet.append(row)


def set_optimal_column_widths(sheet: Worksheet):
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


def make_table(sheet: Worksheet, table_name: str):
    center_aligned_text = NamedStyle(name="center_aligned_text",
                                     alignment=Alignment(horizontal="center", vertical="center"))
    sheet.parent.add_named_style(center_aligned_text)

    table = Table(displayName=table_name, ref=sheet.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True,
                           showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)

    for row in sheet[table.ref]:
        for cell in row:
            cell.style = center_aligned_text


if __name__ == "__main__":
    # Generate a large dataset
    num_rows = 1000
    num_cols = 50
    sample_data = [[f"Data {row}-{col}" for col in range(1, num_cols + 1)] for row in range(1, num_rows + 1)]

    # Create a new workbook and worksheet
    sample_workbook = Workbook()
    sample_worksheet = sample_workbook.active

    # Add data, headers, row index column, and make it a table
    add_data_to_sheet(sample_worksheet, sample_data)
    add_headers(sample_worksheet, [f"Col {i}" for i in range(1, num_cols + 1)])
    add_row_index_column(sample_worksheet)
    make_table(sample_worksheet, "huge_table")
    set_optimal_column_widths(sample_worksheet)

    sample_workbook.save(filename="sample.xlsx")
