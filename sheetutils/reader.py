from typing import List
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook


class SheetReader:
    def __init__(self, filepath: str):
        self.workbook = load_workbook(filepath)
        self.sheet = self.workbook.active

    def get_data(self) -> List[List]:
        read_data = [list(row) for row in self.sheet.iter_rows(values_only=True)]
        return read_data[1:]


if __name__ == "__main__":
    reader = SheetReader(r"C:\Users\Zarei\PycharmProjects\WMan\sample.xlsx")
    data = reader.get_data()
    print(data[0])
