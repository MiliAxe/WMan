from typing import List

from openpyxl.styles import NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


class SheetWriter:
    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active

    def add_row_index_column(self):
        self.sheet.insert_cols(1)
        self.sheet.cell(row=1, column=1, value="Row index")

        for row_num in range(1, self.sheet.max_row):
            self.sheet.cell(row=row_num + 1, column=1, value=str(row_num))

    def add_headers(self, headers: List[str]):
        self.sheet.insert_rows(1)

        for col_index, header in enumerate(headers):
            self.sheet.cell(1, col_index + 1, value=header)

    def add_data_to_sheet(self, data: List[List]) -> None:
        for row in data:
            self.sheet.append(row)

    def set_optimal_column_widths(self):
        for column in self.sheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            self.sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    def make_table(self, table_name: str):
        center_aligned_text = NamedStyle(name="center_aligned_text",
                                         alignment=Alignment(horizontal="center", vertical="center"))
        self.sheet.parent.add_named_style(center_aligned_text)

        table = Table(displayName=table_name, ref=self.sheet.dimensions)
        style = TableStyleInfo(name="TableStyleMedium9",
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        self.sheet.add_table(table)

        for row in self.sheet[table.ref]:
            for cell in row:
                cell.style = center_aligned_text

    def save(self, filename: str):
        self.workbook.save(filename)


if __name__ == "__main__":
    # Generate a large dataset
    num_rows = 1000
    num_cols = 50
    sample_data = [[f"Data {row}-{col}" for col in range(1, num_cols + 1)] for row in range(1, num_rows + 1)]

    writer = SheetWriter()

    # Add data, headers, row index column, and make it a table
    writer.add_data_to_sheet(sample_data)
    writer.add_headers([f"Col {i}" for i in range(1, num_cols + 1)])
    writer.add_row_index_column()
    writer.make_table("huge_table")
    writer.set_optimal_column_widths()

    writer.save(filename="../sample.xlsx")
